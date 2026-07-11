#!/usr/bin/env python3
"""Extract structure and cell data from the COC Excel reference workbook.

This script is intentionally small and deterministic. It does not guess game
rules; it only turns workbook layout into auditable JSON/CSV so data can be
reviewed before being converted into front-end JS files.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import deque
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

DEFAULT_WORKBOOK = Path("books") / "COC7空白卡CY23.3.xlsx"
DEFAULT_OUT_DIR = Path(".tmp_excel_extract")


def cell_to_value(cell: Any) -> Any:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").strip()
    return value


def is_non_empty(value: Any) -> bool:
    return value is not None and value != ""


def safe_name(name: str) -> str:
    value = re.sub(r"[\\/:*?\"<>|\s]+", "_", name.strip())
    return value or "sheet"


def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def range_to_rows(ws: Any, range_ref: str | None = None) -> list[list[Any]]:
    if range_ref:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    else:
        min_row, min_col, max_row, max_col = 1, 1, ws.max_row, ws.max_column
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        rows.append([cell_to_value(cell) for cell in row])
    return rows


def non_empty_cells(ws: Any) -> list[tuple[int, int, Any]]:
    cells: list[tuple[int, int, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell_to_value(cell)
            if is_non_empty(value):
                cells.append((cell.row, cell.column, value))
    return cells


def bounds_from_cells(cells: Iterable[tuple[int, int, Any]]) -> dict[str, Any] | None:
    coords = [(row, col) for row, col, _ in cells]
    if not coords:
        return None
    min_row = min(row for row, _ in coords)
    max_row = max(row for row, _ in coords)
    min_col = min(col for _, col in coords)
    max_col = max(col for _, col in coords)
    return {
        "start": cell_addr(min_row, min_col),
        "end": cell_addr(max_row, max_col),
        "minRow": min_row,
        "maxRow": max_row,
        "minCol": min_col,
        "maxCol": max_col,
    }


def row_previews(ws: Any, max_rows: int) -> list[dict[str, Any]]:
    previews: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        values = [cell_to_value(cell) for cell in row]
        if not any(is_non_empty(value) for value in values):
            continue
        trimmed = list(values)
        while trimmed and not is_non_empty(trimmed[-1]):
            trimmed.pop()
        previews.append({"row": row[0].row, "values": trimmed})
        if len(previews) >= max_rows:
            break
    return previews


def formula_cells(ws: Any, max_cells: int) -> list[dict[str, Any]]:
    formulas: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": value})
                if len(formulas) >= max_cells:
                    return formulas
    return formulas


def merged_ranges(ws: Any) -> list[dict[str, Any]]:
    ranges = []
    for merged in ws.merged_cells.ranges:
        ranges.append({
            "range": str(merged),
            "topLeft": cell_addr(merged.min_row, merged.min_col),
            "value": cell_to_value(ws.cell(merged.min_row, merged.min_col)),
        })
    return ranges


def detect_regions(cells: list[tuple[int, int, Any]], max_regions: int) -> list[dict[str, Any]]:
    remaining = {(row, col) for row, col, _ in cells}
    regions: list[dict[str, Any]] = []
    while remaining and len(regions) < max_regions:
        start = next(iter(remaining))
        queue: deque[tuple[int, int]] = deque([start])
        component: set[tuple[int, int]] = set()
        remaining.remove(start)
        while queue:
            row, col = queue.popleft()
            component.add((row, col))
            for neighbor in ((row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1)):
                if neighbor in remaining:
                    remaining.remove(neighbor)
                    queue.append(neighbor)
        min_row = min(row for row, _ in component)
        max_row = max(row for row, _ in component)
        min_col = min(col for _, col in component)
        max_col = max(col for _, col in component)
        regions.append({
            "range": f"{cell_addr(min_row, min_col)}:{cell_addr(max_row, max_col)}",
            "cellCount": len(component),
            "minRow": min_row,
            "maxRow": max_row,
            "minCol": min_col,
            "maxCol": max_col,
        })
    regions.sort(key=lambda item: (item["minRow"], item["minCol"]))
    return regions


def build_sheet_map(ws: Any, max_preview_rows: int, max_formulas: int, max_regions: int) -> dict[str, Any]:
    cells = non_empty_cells(ws)
    return {
        "name": ws.title,
        "dimensions": ws.calculate_dimension(),
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "nonEmptyCount": len(cells),
        "nonEmptyBounds": bounds_from_cells(cells),
        "mergedRanges": merged_ranges(ws),
        "formulaCells": formula_cells(ws, max_formulas),
        "detectedRegions": detect_regions(cells, max_regions),
        "rowPreviews": row_previews(ws, max_preview_rows),
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def command_map(args: argparse.Namespace) -> None:
    workbook_path = Path(args.workbook)
    out_dir = Path(args.out_dir)
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    payload = {
        "workbook": str(workbook_path),
        "sheetNames": wb.sheetnames,
        "sheets": [build_sheet_map(ws, args.max_preview_rows, args.max_formulas, args.max_regions) for ws in wb.worksheets],
    }
    write_json(out_dir / "workbook_map.json", payload)
    print(f"wrote {out_dir / 'workbook_map.json'}")


def command_dump(args: argparse.Namespace) -> None:
    workbook_path = Path(args.workbook)
    out_dir = Path(args.out_dir)
    wb = load_workbook(workbook_path, data_only=args.data_only, read_only=False)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"sheet not found: {args.sheet}. Available: {', '.join(wb.sheetnames)}")
    ws = wb[args.sheet]
    rows = range_to_rows(ws, args.range)
    base = safe_name(args.sheet)
    suffix = args.range.replace(":", "_") if args.range else "full"
    payload = {
        "workbook": str(workbook_path),
        "sheet": args.sheet,
        "range": args.range or ws.calculate_dimension(),
        "dataOnly": args.data_only,
        "rows": rows,
    }
    write_json(out_dir / f"{base}_{suffix}.json", payload)
    write_csv(out_dir / f"{base}_{suffix}.csv", rows)
    print(f"wrote {out_dir / f'{base}_{suffix}.json'}")
    print(f"wrote {out_dir / f'{base}_{suffix}.csv'}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract map/data from the COC Excel workbook.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to the .xlsx workbook.")
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR), help="Directory for generated JSON/CSV files.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    map_parser = subparsers.add_parser("map", help="Write workbook_map.json with sheet structure previews.")
    map_parser.add_argument("--max-preview-rows", type=int, default=40)
    map_parser.add_argument("--max-formulas", type=int, default=80)
    map_parser.add_argument("--max-regions", type=int, default=80)
    map_parser.set_defaults(func=command_map)

    dump_parser = subparsers.add_parser("dump", help="Dump one sheet or range to JSON and CSV.")
    dump_parser.add_argument("--sheet", required=True, help="Sheet name to dump.")
    dump_parser.add_argument("--range", help="Optional A1 range, such as A1:K30.")
    dump_parser.add_argument("--data-only", action="store_true", help="Read cached formula values instead of formulas.")
    dump_parser.set_defaults(func=command_dump)
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
